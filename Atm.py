
import os
from openpyxl import Workbook, load_workbook

# Excel file name
FILE_NAME = "atm_accounts.xlsx"


# ==========================================
# CREATE EXCEL FILE AUTOMATICALLY
# ==========================================
def create_excel_file():

    if not os.path.exists(FILE_NAME):

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Accounts"

        # Column headings
        sheet.append([
            "Account Number",
            "Name",
            "PIN",
            "Balance"
        ])

        workbook.save(FILE_NAME)

        print("Excel file created successfully!")


# ==========================================
# CREATE ACCOUNT
# ==========================================
def create_account():

    workbook = load_workbook(FILE_NAME)
    sheet = workbook["Accounts"]

    print("\n----- CREATE ACCOUNT -----")

    name = input("Enter your name: ")
    pin = input("Create a 4-digit PIN: ")

    # Check PIN
    if len(pin) != 4 or not pin.isdigit():
        print("PIN must contain exactly 4 digits.")
        workbook.close()
        return

    # Generate account number
    if sheet.max_row == 1:
        account_number = 1001
    else:
        last_account = sheet.cell(sheet.max_row, 1).value
        account_number = int(last_account) + 1

    # Initial balance
    balance = 0

    # Add account to Excel
    sheet.append([
        account_number,
        name,
        pin,
        balance
    ])

    workbook.save(FILE_NAME)
    workbook.close()

    print("\nAccount created successfully!")
    print("Your Account Number:", account_number)
    print("Your Initial Balance: ₹0")


# ==========================================
# LOGIN
# ==========================================
def login():

    account_number = input("\nEnter Account Number: ")
    pin = input("Enter PIN: ")

    try:
        account_number = int(account_number)
    except ValueError:
        print("Invalid account number.")
        return None, None, None

    workbook = load_workbook(FILE_NAME)
    sheet = workbook["Accounts"]

    # Search account
    for row in range(2, sheet.max_row + 1):

        stored_account = sheet.cell(row, 1).value
        stored_pin = str(sheet.cell(row, 3).value)

        if stored_account == account_number and stored_pin == pin:

            print("Login successful!")

            return workbook, sheet, row

    workbook.close()

    print("Invalid Account Number or PIN.")

    return None, None, None


# ==========================================
# DEPOSIT MONEY
# ==========================================
def deposit_money():

    workbook, sheet, row = login()

    if row is None:
        return

    try:

        amount = float(input("Enter amount to deposit: "))

        if amount <= 0:
            print("Please enter a valid amount.")
            workbook.close()
            return

        current_balance = float(sheet.cell(row, 4).value)

        new_balance = current_balance + amount

        sheet.cell(row, 4).value = new_balance

        workbook.save(FILE_NAME)
        workbook.close()

        print("\nDeposit successful!")
        print("Deposited Amount: ₹", amount)
        print("New Balance: ₹", new_balance)

    except ValueError:

        print("Please enter a valid number.")
        workbook.close()


# ==========================================
# WITHDRAW MONEY
# ==========================================
def withdraw_money():

    workbook, sheet, row = login()

    if row is None:
        return

    try:

        amount = float(input("Enter amount to withdraw: "))

        if amount <= 0:
            print("Please enter a valid amount.")
            workbook.close()
            return

        current_balance = float(sheet.cell(row, 4).value)

        # Check balance
        if amount > current_balance:

            print("\nInsufficient balance!")

        else:

            new_balance = current_balance - amount

            sheet.cell(row, 4).value = new_balance

            workbook.save(FILE_NAME)

            print("\nWithdrawal successful!")
            print("Withdrawn Amount: ₹", amount)
            print("Remaining Balance: ₹", new_balance)

    except ValueError:

        print("Please enter a valid number.")

    workbook.close()


# ==========================================
# CHECK BALANCE
# ==========================================
def check_balance():

    workbook, sheet, row = login()

    if row is None:
        return

    balance = sheet.cell(row, 4).value

    print("\n----- ACCOUNT BALANCE -----")
    print("Your Balance: ₹", balance)

    workbook.close()


# ==========================================
# ACCOUNT DETAILS
# ==========================================
def account_details():

    workbook, sheet, row = login()

    if row is None:
        return

    account_number = sheet.cell(row, 1).value
    name = sheet.cell(row, 2).value
    balance = sheet.cell(row, 4).value

    print("\n----- ACCOUNT DETAILS -----")
    print("Account Number:", account_number)
    print("Name:", name)
    print("Balance: ₹", balance)

    workbook.close()


# ==========================================
# MAIN ATM MENU
# ==========================================
def atm_menu():

    while True:

        print("\n")
        print("==============================")
        print("        ATM MACHINE")
        print("==============================")
        print("1. Create Account")
        print("2. Deposit Money")
        print("3. Withdraw Money")
        print("4. Check Balance")
        print("5. Account Details")
        print("6. Exit")
        print("==============================")

        choice = input("Enter your choice: ")

        if choice == "1":

            create_account()

        elif choice == "2":

            deposit_money()

        elif choice == "3":

            withdraw_money()

        elif choice == "4":

            check_balance()

        elif choice == "5":

            account_details()

        elif choice == "6":

            print("\nThank you for using the ATM!")
            print("Have a nice day!")
            break

        else:

            print("\nInvalid choice!")
            print("Please select a number from 1 to 6.")


# ==========================================
# PROGRAM START
# ==========================================

create_excel_file()

atm_menu()
